import openpyxl

book = openpyxl.load_workbook("C:\\Users\\newaj\\Documents\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
Dict = {}   #{'Name': 'TC2', 'FirstName': 'b', 'LastName': 'g', 'Gender': 'l'}
print(cell.value)
sheet.cell(row=2,column=2).value = "Shah Newaj"
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for i in range(1,sheet.max_row+1):  #to get rows
    if(sheet.cell(row=i,column=1).value) == "TC2":
        for j in range(1, sheet.max_column + 1):    #to get columns
            # {'Name': 'TC2', 'FirstName': 'b', 'LastName': 'g', 'Gender': 'l'}
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)