import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome()

def update_excel_data(filePath, searchTerm, colName, new_Value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    # cell = sheet.cell(row=1,column=2)
    Dict = {}   #{'Name': 'TC2', 'FirstName': 'b', 'LastName': 'g', 'Gender': 'l'}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == colName:
            Dict["col"] = i

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_Value
    book.save(file_path)

# Global Time out - max 5 seconds
driver.implicitly_wait(5)   #Applying Implicit Wait
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.maximize_window()

# file_path = "C:\\Users\\newaj\\Downloads\\download.xlsx"
file_path = "C:/Users/newaj/Downloads/download.xlsx"
fruit_name = "Apple"
newValue = "999"

time.sleep(2)
driver.find_element(By.ID,"downloadButton").click()
time.sleep(3)

# edit excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)


#------------------Upload---------------------
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver,5)     #initializing Explicit wait
# XPATH = //div[@class='Toastify__toast-body']/div[2]
toast_locator = (By.XPATH,"//div[@class='Toastify__toast-body']/div[2]")    #CSS = .Toastify__toast-body div:nth-child(2)
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
#----------------------------------- Dynamic Locator: Smart XPATH -------------------------------------------------------------------------
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
#-------------------------------------------------------------------------------------------------------------------------------
assert actual_price == newValue